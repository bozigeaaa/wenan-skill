import openpyxl
import os

def dump_xlsx_to_txt(xlsx_path, txt_path):
    print(f"Dumping {xlsx_path} to {txt_path}...")
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        with open(txt_path, 'w', encoding='utf-8') as f:
            f.write(f"Workbook: {os.path.basename(xlsx_path)}\n\n")
            for sheet_name in wb.sheetnames:
                f.write(f"=== Sheet: {sheet_name} ===\n")
                sheet = wb[sheet_name]
                for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                    # Join row values as string, handle None
                    row_str = " | ".join([str(val).replace('\n', ' ') if val is not None else "" for val in row])
                    # If the row is completely empty, skip it
                    if not row_str.strip().replace('|', ''):
                        continue
                    f.write(f"L{r_idx+1}: {row_str}\n")
                f.write("\n")
        print(f"Done dumping {xlsx_path}.")
    except Exception as e:
        print("Error:", e)

dump_xlsx_to_txt("knowledge/沙特临建行业认知.xlsx", "knowledge/沙特临建行业认知_dump.txt")
dump_xlsx_to_txt("knowledge/钧瀚产品优势分级分类总表_v4.xlsx", "knowledge/钧瀚产品优势分级分类总表_dump.txt")
