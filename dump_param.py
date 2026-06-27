import openpyxl

wb = openpyxl.load_workbook("knowledge/K系列卫生间参数.xlsx", data_only=True)
print("Sheets:", wb.sheetnames)
with open("knowledge/K系列卫生间参数_dump.txt", 'w', encoding='utf-8') as f:
    for sheet_name in wb.sheetnames:
        f.write(f"=== Sheet: {sheet_name} ===\n")
        sheet = wb[sheet_name]
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            row_str = " | ".join([str(val).replace('\n', ' ') if val is not None else "" for val in row])
            if row_str.strip().replace('|', ''):
                f.write(f"L{r_idx+1}: {row_str}\n")
        f.write("\n")
print("Done dumping parameters sheet.")
