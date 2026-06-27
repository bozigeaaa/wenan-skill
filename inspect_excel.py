import openpyxl

def inspect_xlsx(path):
    print(f"Inspecting {path}:")
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        print("Sheets:", wb.sheetnames)
        for name in wb.sheetnames[:3]:
            sheet = wb[name]
            print(f"\nSheet: {name}")
            # Print first 5 rows
            for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if r_idx >= 5:
                    break
                print(row)
    except Exception as e:
        print("Error:", e)

inspect_xlsx("knowledge/沙特临建行业认知.xlsx")
print("="*40)
inspect_xlsx("knowledge/钧瀚产品优势分级分类总表_v4.xlsx")
